#!/usr/bin/python3
# Geoff Dillon geoff_dillon@dell.com
# Copyright Dell, Inc 2024
# FOR INTERNAL USE ONLY.  DO NOT distribute to customers or partners/vendors.
# Python3 script to compare Excel task trackers for Coreweave internal snd external management

import os
import sys
import openpyxl
import subprocess
import argparse
import base64
import configparser
import datetime
from dataclasses import dataclass

@dataclass
class Issue:
    id: str
    row: int
    age: int
    opendate: datetime.datetime
    closedate: datetime.datetime
    requester: str
    owner: str
    state: str
    priority: str
    temperature: str
    issuetype: str
    needby: datetime.datetime
    dell_tkt: str
    cw_tkt: str
    ami_tkt: str
    nv_tkt: str
    platform: str
    intext: str
    eta: str  # can't trust this to be a datetime
    description: str
    ismatched: bool  # will save time in diffing
    
    
print_verbose = False
check_descriptions = False

#does not include comment fields
intheads = {
    'ID': 0,
    'AGE': 1,
    'OPENDATE':2,
    'CLOSEDATE':3,
    'REQUESTER':4,
    'OWNER':5,
    'STATE': 6,
    'PRIORITY':7,
    'TEMP':8,
    'ISSUETYPE':9,
    'NEEDBY':10,
    'DELL_TKT':11,
    'CW_TKT':12,
    'AMI_TKT':13,
    'NV_TKT':14,
    'PLATFORM':15,
    'INTEXT':16,
    'DESCRIPTION':17,
}

extheads = {
    'ID': 0,
    'AGE': 1,
    'OPENDATE':2,
    'CLOSEDATE':3,
    'REQUESTER':4,
    'OWNER':5,
    'STATE': 6,
    'PRIORITY':7,
    'TEMP':8,
    'ISSUETYPE':9,
    'NEEDBY':10,
    'DELL_TKT':11,
    'CW_TKT':12,
    'AMI_TKT':13,
    'NV_TKT':14,
    'PLATFORM':15,
    'DESCRIPTION':16,
    'ETA': 17,
}

int_issues = []
ext_issues = []

def verbose(*args):
    if print_verbose:
        for arg in args:
            print(arg)
            
def reporterror(*args):
    print(f"ERROR: {args}")

def findissuebyID(issueslist, id):
    for issue in issueslist:
        if (issue.id == id):
            return issue
    return None
    
    
if __name__ == "__main__":
    PARSER = argparse.ArgumentParser(description=__doc__)
    PARSER.add_argument('-v', '--verbose', action='store_true', default=False, help="Print more messages.")
    PARSER.add_argument('-d', '--descriptions', action='store_true', default=False, help="Check for matching descriptions.")
    PARSER.add_argument('-i', '--internal-tracker', type=str, required=True, help="The filename and path of the Internal Tracker Excel Spreadsheet (.XLSX)")
    PARSER.add_argument('-e', '--external-tracker', type=str, required=True, help="The filename and path of the Esternal Tracker Excel Spreadsheet (.XLSX)")
    
    args = PARSER.parse_args()

    if (sys.version_info.major < 3):
        print("This script requires Python version 3 or higher.  You are running {}.{}".format(sys.version_info.major, sys.version_info.minor))
        sys.exit(1)
    # verify file format is XLSX by extension
    if (not '.xlsx' in args.internal_tracker.lower()):
        print("The internal tracker must be in XLSX format.")
        sys.exit(1)
    if (not '.xlsx' in args.external_tracker.lower()):
        print("The external tracker must be in XLSX format.")
        sys.exit(1)
    # verify files exist
    if (not os.path.isfile(args.internal_tracker)):
        print("The internal tracker file path does not exist.")
        sys.exit(1)
    if  (not os.path.isfile(args.external_tracker)):
        print("The external tracker file path does not exist.")
        sys.exit(1)
    if (args.verbose):
        print_verbose = True
    if (args.descriptions):
        check_descriptions = True
        
    # verify that the workbooks can be loaded
    try:
        int_wb = openpyxl.load_workbook(filename=args.internal_tracker, read_only=True, data_only=True)
        verbose("Internal WB Sheets:", int_wb.sheetnames)
    except:
        print("Problem loading the excel workbook ", args,internal_tracker)
        exit(1)
    try:
        ext_wb = openpyxl.load_workbook(filename=args.external_tracker, read_only=True, data_only=True)
        verbose("External WB Sheets:", ext_wb.sheetnames)
    except:
        print("Problem loading the excel workbook ", args,external_tracker)
        exit(1)
        
    int_wb_sheet = int_wb.active
    verbose("Internal WB Active Sheet title ", int_wb_sheet.title)
    
    ext_wb_sheet = ext_wb.active
    verbose("External WB Active Sheet title ", ext_wb_sheet.title)
    # ingest the internal issues list
    print("Processing internal issues list")
    rowidx=2
    for introw in int_wb_sheet.iter_rows(min_row=2, values_only=True):
        if (introw[intheads['ID']] == None):
            reporterror(f"NO ID: Skipping the issue at row {rowidx} because id is blank.")
        elif (introw[intheads['STATE']] not in ('open', 'closed')): 
            reporterror(f"INVALID STATE: Skipping the issue at row {rowidx} with id {introw[intheads['ID']]} because state is invalid: {introw[intheads['STATE']]}")
        else:
            verbose(f"Creating issue at row {rowidx} with id {introw[intheads['ID']]} and state {introw[intheads['STATE']]}")
            existissue = findissuebyID(int_issues, introw[intheads['ID']])
            if (existissue):
                reporterror(f"DUPLICATE ID: The id {introw[intheads['ID']]} at row {rowidx} is already in the list from row {existissue.row}")
            if (introw[intheads['AGE']] == None):
                reporterror(f"INVALID AGE: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank Age field.")
            if (introw[intheads['OPENDATE']] == None):
                reporterror(f"INVALID OPENDATE: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank Open Date field.")
            if (introw[intheads['STATE']] == 'open'):
                if (introw[intheads['PRIORITY']] == None):
                    reporterror(f"INVALID PRIORITY: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank priority field.")
                if (introw[intheads['TEMP']] == None):
                    reporterror(f"INVALID TEMP: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank Temperature field.")
                if (introw[intheads['PLATFORM']] == None):
                    reporterror(f"INVALID PLATFORM: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank platform field.")
                if (introw[intheads['DESCRIPTION']] == None):
                    reporterror(f"INVALID DESCRIPTION: The issue at row {rowidx} with id {introw[intheads['ID']]} has a blank description field.")

            anissue = Issue(id=introw[intheads['ID']],
                row=rowidx,
                age=introw[intheads['AGE']],
                opendate=introw[intheads['OPENDATE']],
                closedate=introw[intheads['CLOSEDATE']],
                requester=introw[intheads['REQUESTER']],
                owner=introw[intheads['OWNER']],
                state=introw[intheads['STATE']],
                priority=introw[intheads['PRIORITY']],
                temperature=introw[intheads['TEMP']],
                issuetype=introw[intheads['ISSUETYPE']],
                needby=introw[intheads['NEEDBY']],
                dell_tkt=introw[intheads['DELL_TKT']],
                cw_tkt=introw[intheads['CW_TKT']],
                ami_tkt=introw[intheads['AMI_TKT']],
                nv_tkt=introw[intheads['NV_TKT']],
                platform=introw[intheads['PLATFORM']],
                intext=introw[intheads['INTEXT']],
                eta=None,
                description=introw[intheads['DESCRIPTION']],
                ismatched=False                           
            )
            #verbose(anissue)
            int_issues.append(anissue)

        rowidx += 1
        
    # ingest the external issues list
    print("Processing external issues list")
    rowidx=2
    for row in ext_wb_sheet.iter_rows(min_row=2, values_only=True):
        if (row[extheads['ID']] == None):
            reporterror(f"NO ID: Skipping the issue at row {rowidx} because id is blank.")
        elif (row[extheads['STATE']] not in ('open', 'closed')): 
            reporterror(f"INVALID STATE: Skipping the issue at row {rowidx} with id {row[extheads['ID']]} because state is invalid: {row[extheads['STATE']]}")
        else:
            verbose(f"Creating issue at row {rowidx} with id {row[extheads['ID']]} and state {row[extheads['STATE']]}")
            existissue = findissuebyID(ext_issues, row[extheads['ID']])
            if (existissue):
                reporterror(f"DUPLICATE ID: The id {row[extheads['ID']]} at row {rowidx} is already in the list from row {existissue.row}")
            if (row[extheads['AGE']] == None):
                reporterror(f"INVALID AGE: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank Age field.")
            if (row[extheads['OPENDATE']] == None):
                reporterror(f"INVALID OPENDATE: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank Open Date field.")
            if (row[extheads['STATE']] == 'open'):
                if (row[extheads['PRIORITY']] == None):
                    reporterror(f"INVALID PRIORITY: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank priority field.")
                if (row[extheads['TEMP']] == None):
                    reporterror(f"INVALID TEMP: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank Temperature field.")
                if (row[extheads['PLATFORM']] == None):
                    reporterror(f"INVALID PLATFORM: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank platform field.")
                if (row[extheads['DESCRIPTION']] == None):
                    reporterror(f"INVALID DESCRIPTION: The issue at row {rowidx} with id {row[extheads['ID']]} has a blank description field.")
    
            anissue = Issue(id=row[extheads['ID']],
                row=rowidx,
                age=row[extheads['AGE']],
                opendate=row[extheads['OPENDATE']],
                closedate=row[extheads['CLOSEDATE']],
                requester=row[extheads['REQUESTER']],
                owner=row[extheads['OWNER']],
                state=row[extheads['STATE']],
                priority=row[extheads['PRIORITY']],
                temperature=row[extheads['TEMP']],
                issuetype=row[extheads['ISSUETYPE']],
                needby=row[extheads['NEEDBY']],
                dell_tkt=row[extheads['DELL_TKT']],
                cw_tkt=row[extheads['CW_TKT']],
                ami_tkt=row[extheads['AMI_TKT']],
                nv_tkt=row[extheads['NV_TKT']],
                platform=row[extheads['PLATFORM']],
                intext=None,
                eta=row[extheads['ETA']],
                description=row[extheads['DESCRIPTION']],
                ismatched=False                           
            )
            #verbose(anissue)
            ext_issues.append(anissue)

        rowidx += 1
    
    print("Checking for Internal/External differences")
    
    for issue in ext_issues:
        in_issue = findissuebyID(int_issues, issue.id)
        if (not in_issue):
            reporterror(f"EXT ONLY ISSUE: The id {issue.id} is not copied to the internal issues list.")
        else:
            if (in_issue.state != issue.state):
                reporterror(f"STATE MISMATCH: The id {issue.id} has different states. internal = {in_issue.state} | external = {issue.state}.")
            if (in_issue.priority != issue.priority):
                reporterror(f"PRIORITY MISMATCH: The id {issue.id} has different priorities. internal = {in_issue.priority} | external = {issue.priority}.")
            if (in_issue.temperature != issue.temperature):
                reporterror(f"TEMP MISMATCH: The id {issue.id} has different temperatures. internal = {in_issue.temperature} | external = {issue.temperature}.")
            if (check_descriptions and (in_issue.state == 'open') and (in_issue.description != issue.description)):
                reporterror(f"DESC MISMATCH: The id {issue.id} has different descriptions. internal = {in_issue.description} | external = {issue.description}.")
    for in_issue in int_issues:
        if (in_issue.intext and (in_issue.intext.lower() != 'internal')):
            ext_issue = findissuebyID(ext_issues, in_issue.id)
            if (not ext_issue):
                reporterror(f"INT ONLY ISSUE: The id {in_issue.id} at row {in_issue.row} is not marked as Internal and is not copied to the external isues list.")
                
    sys.exit(0)

